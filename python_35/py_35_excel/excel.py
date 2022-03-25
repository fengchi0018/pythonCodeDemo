"""
======================
Time:2022/3/15 14:21
Author:feng
========================
"""
import openpyxl

# 加载excel文件
# workbook = openpyxl.load_workbook("测试用例.xlsx")
# # 获取表单名
# print(workbook.sheetnames)
# # 选中表单
# sh = workbook["Sheet1"]
# print(sh)
# # 读取数据
# print(sh.cell(row=1, column=1).value)
# print(sh.cell(column=2, row=1).value)


# -----------读取表单中所有数据--------------

workbook = openpyxl.load_workbook("测试用例.xlsx")
sh = workbook["Sheet1"]
# rows：按行获取表单中所有的格子，每一行的格子放到一个元组中
res = list(sh.rows)
# 获取excel中第一行的数据
title = [i.value for i in res[0]]
cases = []
# 遍历第一行以外所有的数据
for item in res[1:]:
    # 获取改行数据
    data = [i.value for i in item]
    # 两个列表【title,data]打包成字典
    dic = dict(zip(title, data))
    # 字典添加至列表中
    cases.append(dic)
print("-----cases---:",cases)
