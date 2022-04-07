"""
@Time    : 2022/4/6 20:20
@Author  : feng
"""
import openpyxl


class HandExcel:
    def __init__(self, filename, sheetname):
        self.filname = filename
        self.sheetname = sheetname

    def read_data(self):
        workbook = openpyxl.load_workbook(self.filname)
        sh = workbook[self.sheetname]
        res = list(sh.rows)
        title = [i.value for i in res[0]]
        cases = []
        for item in res[1:]:
            data = [i.value for i in item]
            case = dict(zip(title, data))
            cases.append(case)
        return cases

    def write_data(self, row, column, value):
        workbook = openpyxl.load_workbook(self.filname)
        sh = workbook[self.sheetname]
        sh.cell(row=row, column=column, value=value)
        workbook.save(self.filname)


if __name__ == '__main__':
    book = HandExcel(r"D:\PycharmProjects\Python_C\python_35\demo\datas\测试用例.xlsx", "Sheet1")
    book.write_data(2, 1, "测试")

